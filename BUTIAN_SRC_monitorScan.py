import requests
import openpyxl
import pymysql
import json
import random
import time
import yagmail
from requests.packages import urllib3
from bs4 import BeautifulSoup

# 该脚本可扫描补天的指定范围内的企业SRC（本版本不支持扫描企业SRC）；同时扫描结果和数据库对比，新增的数据会通过邮件附件的xlsx发送。

USER_AGENTS = [
    "Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; SV1; AcooBrowser; .NET CLR 1.1.4322; .NET CLR 2.0.50727)",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 6.0; Acoo Browser; SLCC1; .NET CLR 2.0.50727; Media Center PC 5.0; .NET CLR 3.0.04506)",
    "Mozilla/4.0 (compatible; MSIE 7.0; AOL 9.5; AOLBuild 4337.35; Windows NT 5.1; .NET CLR 1.1.4322; .NET CLR 2.0.50727)",
    "Mozilla/5.0 (Windows; U; MSIE 9.0; Windows NT 9.0; en-US)",
    "Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Win64; x64; Trident/5.0; .NET CLR 3.5.30729; .NET CLR 3.0.30729; .NET CLR 2.0.50727; Media Center PC 6.0)",
    "Mozilla/5.0 (compatible; MSIE 8.0; Windows NT 6.0; Trident/4.0; WOW64; Trident/4.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; .NET CLR 1.0.3705; .NET CLR 1.1.4322)",
    "Mozilla/4.0 (compatible; MSIE 7.0b; Windows NT 5.2; .NET CLR 1.1.4322; .NET CLR 2.0.50727; InfoPath.2; .NET CLR 3.0.04506.30)",
    "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN) AppleWebKit/523.15 (KHTML, like Gecko, Safari/419.3) Arora/0.3 (Change: 287 c9dfb30)",
    "Mozilla/5.0 (X11; U; Linux; en-US) AppleWebKit/527+ (KHTML, like Gecko, Safari/419.3) Arora/0.6",
    "Mozilla/5.0 (Windows; U; Windows NT 5.1; en-US; rv:1.8.1.2pre) Gecko/20070215 K-Ninja/2.1.1",
    "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN; rv:1.9) Gecko/20080705 Firefox/3.0 Kapiko/3.0",
    "Mozilla/5.0 (X11; Linux i686; U;) Gecko/20070322 Kazehakase/0.4.5",
    "Mozilla/5.0 (X11; U; Linux i686; en-US; rv:1.9.0.8) Gecko Fedora/1.9.0.8-1.fc10 Kazehakase/0.5.6",
    "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.56 Safari/535.11",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_7_3) AppleWebKit/535.20 (KHTML, like Gecko) Chrome/19.0.1036.7 Safari/535.20",
    "Opera/9.80 (Macintosh; Intel Mac OS X 10.6.8; U; fr) Presto/2.9.168 Version/11.52",
    "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.11 (KHTML, like Gecko) Chrome/20.0.1132.11 TaoBrowser/2.0 Safari/536.11",
    "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/21.0.1180.71 Safari/537.1 LBBROWSER",
    "Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; WOW64; Trident/5.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; .NET4.0C; .NET4.0E; LBBROWSER)",
    "Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; SV1; QQDownload 732; .NET4.0C; .NET4.0E; LBBROWSER)",
    "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.84 Safari/535.11 LBBROWSER",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 6.1; WOW64; Trident/5.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; .NET4.0C; .NET4.0E)",
    "Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; WOW64; Trident/5.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; .NET4.0C; .NET4.0E; QQBrowser/7.0.3698.400)",
    "Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; SV1; QQDownload 732; .NET4.0C; .NET4.0E)",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; Trident/4.0; SV1; QQDownload 732; .NET4.0C; .NET4.0E; 360SE)",
    "Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; SV1; QQDownload 732; .NET4.0C; .NET4.0E)",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 6.1; WOW64; Trident/5.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; .NET4.0C; .NET4.0E)",
    "Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/21.0.1180.89 Safari/537.1",
    "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/21.0.1180.89 Safari/537.1",
    "Mozilla/5.0 (iPad; U; CPU OS 4_2_1 like Mac OS X; zh-cn) AppleWebKit/533.17.9 (KHTML, like Gecko) Version/5.0.2 Mobile/8C148 Safari/6533.18.5",
    "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:2.0b13pre) Gecko/20110307 Firefox/4.0b13pre",
    "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:16.0) Gecko/20100101 Firefox/16.0",
    "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.11 (KHTML, like Gecko) Chrome/23.0.1271.64 Safari/537.11",
    "Mozilla/5.0 (X11; U; Linux x86_64; zh-CN; rv:1.9.2.10) Gecko/20100922 Ubuntu/10.10 (maverick) Firefox/3.6.10"
]
headers = {
    'User-Agent': USER_AGENTS[random.randint(0,33)],
    'Content-Type': "application/x-www-form-urlencoded",
}

# 自动邮件脚本
class emailAuto:
    def __init__(self, user='****@163.com', password='****',  host='smtp.163.com'):
        self.user = user
        self.password = password
        self.host = host
        self.emils = yagmail.SMTP(user=self.user, password=self.password, host=self.host)
        self.mailData()
    # 邮件的内容数据
    def mailData(self):
        self.mailDatas = {
            'title': '补天SRC变动监控_'  + str(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())),
            'contents': '',
            'received': {
                'to': ['***@qq.com','****@tom.com'],
                'cc': [],
                'bc': [],
            },
            'atta': [],
        }
    # 日志输出，当前目录下的sending.log
    def autoLog(self):
        logFile = open('./sending.log' ,'a+')
        # 时间 - 主题 - 收件人 - 抄送人 - 发件人 - 附件 - 正文
        log = "【%s】 - %s - %s — %s - %s - %s - %s \r\n"%(
            time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()),
            self.mailDatas['title'],
            self.mailDatas['received']['to'],
            self.mailDatas['received']['cc'],
            self.user,
            self.mailDatas['atta'],
            self.mailDatas['contents']
        )
        logFile.write(log)
        logFile.close()
    # 按html发送数据
    def htmlForm(self, addr):
        self.addr = addr
        self.htmlText = open(self.addr, "r").read()
    # 发送邮件
    def sending(self):
        self.data = self.mailDatas
        try:
            self.emils.send(
                to=self.data['received']['to'],
                cc=self.data['received']['cc'],
                bcc=self.data['received']['bc'],
                subject=self.data['title'],
                contents=self.data['contents'],
                attachments=self.data['atta'],
            )
            print("###### Email sent successfully ######")
            self.autoLog()
        except :
            print('邮件发生错误')

def company():
    urls = 'https://www.butian.net/Company/'
    data = []
    for i in range(63000, 70000):
        url = urls + str(i)
        try:
            res = requests.post(url, headers=headers, allow_redirects=False, verify=False)
        except requests.exceptions.ConnectionError:
            print("网络错误,连接中止.")
            exit()
        except requests.exceptions.HTTPError:
            continue
        if (res.status_code == 200):
            # print("{0}\t{1}".format(res.url, res.status_code))
            texts = BeautifulSoup(res.text, 'html.parser')
            name = texts.find(class_='firmName1').get_text()  # 企业名称

            spp1, spp2 = texts.find_all(class_='spp')
            spp1, spp2 = spp1.text, spp2.text

            firmDec = texts.find(class_='firmDec')
            firmDec.dl.decompose()
            firmDec.dl.decompose()
            # ?-该处的解析未解决，无法按要求输出注释内容。
            firmDec = str(firmDec)

            list = {
                'company_id': i,
                'name': name,  # 企业名称
                'company_type': firmDec[-13:-11],  # 公益/专属
                'spp1': spp1,  # 已发现漏洞数/最高赏金
                'spp2': spp2,  # 已解决漏洞数/累计赏金
                'urls': res.url,  # src_url
            }
            # print(list)
            data.append(list)
        else:
            continue
    # print(data)
    company = {
        'status': 'ok',
        'length': len(data),
        'data': data,
    }
    # print(company)
    # print(json.dumps(company))
    return json.dumps(company)

def sqlSave(data):
    db_con = pymysql.connections.Connection(
        host="175.**.**.**",
        user="****",
        password="123456",
        database="API")
    cursor = db_con.cursor()

    sql_select = 'SELECT BT_src.company_id FROM BT_src ORDER BY BT_src.company_id ASC;'
    sql_update = "UPDATE `BT_src` SET `name`='{name}', `company_id`='{company_id}', `company_type`='{company_type}', `spp1`='{spp1}', `spp2`='{spp2}', `urls`='{urls}', `update_time`='{time}' WHERE (`company_id`='{company_id}' ) ;"
    sql_insert = "INSERT INTO `BT_src` (`name`, `company_id`, `company_type`, `spp1`, `spp2`, `urls`, `find_time`, `update_time`) VALUES ('{name}','{company_id}','{company_type}','{spp1}','{spp2}','{urls}','{time}','{time}') ;"

    cursor.execute(sql_select)
    result = cursor.fetchall()

    id_list = []
    for d in result:
        id_list.append(d[0])

    company = json.loads(data)
    datas = company['data']
    lens = {"insert": 0, "update": 0}
    insertDate = []
    for data in datas:
        try:
            if data['company_id'] in id_list:
                data['time'] = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
                cursor.execute(sql_update.format(**data))
                db_con.commit()
                lens['update'] = lens['update'] + 1
            else:
                insertDate.append(data)
                data['time'] = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
                cursor.execute(sql_insert.format(**data))
                db_con.commit()
                lens['insert'] = lens['insert'] + 1
        except:
            result = "数据库数据更新/插入发生错误，请及时处理。"
            emailAutos = emailAuto()
            emailAutos.mailDatas['title'] = '补天SRC变动监控脚本发生错误'
            emailAutos.mailDatas['contents'] = result
            emailAutos.sending()

    result = "本次共计更新 {update} 条数据，新增 {insert} 条数据。".format(**lens)
    return result, insertDate
    # print(insertDate)
# {"company_id": 63350, "name": "", "company_type": "", "spp1": "0", "spp2": "0", "urls": ""}
def companyOut(data):
    # print(data)
    Out = "| {0:^6} | {5:<40} | {2:{6}^4} | {3:^10} | {4:^10} | {1:<10}"
    wb = openpyxl.Workbook()
    ws = wb.active

    sheetTitleList = ['id','企业名称','类型','已发现/最高赏金','已解决/累计赏金','SRC地址']
    k = 1
    for titles in sheetTitleList:
        ws.cell(row=1, column=k, value=titles)
        k = k+1
    row = 2
    for i in data:
        ws.cell(row=row, column=1, value=i['company_id'])
        ws.cell(row=row, column=2, value=i['name'])
        ws.cell(row=row, column=3, value=i['company_type'])
        ws.cell(row=row, column=4, value=i['spp1'])
        ws.cell(row=row, column=5, value=i['spp2'])
        ws.cell(row=row, column=6, value=i['urls'])
        row = row+1
    fileName = time.strftime("%Y%m%d%H%M%S", time.localtime())
    wb.save(fileName + '.xlsx')

    return fileName


if __name__ == '__main__':
    urllib3.disable_warnings()
    emailAutos = emailAuto()
    data = company()
    result, insertDate = sqlSave(data)
    if len(insertDate):
        emailAutos.mailDatas['contents'] = result + "请查看附件。"
        fileName = companyOut(insertDate)
        emailAutos.mailDatas['atta'] = fileName + ".xlsx"
    else:
        emailAutos.mailDatas['contents'] = result + "本次无新增SRC企业。"

    emailAutos.sending()


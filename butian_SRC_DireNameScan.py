import requests
from random import choice
import argparse # 用于命令选项解析
import json
import openpyxl
import random
from requests.packages import urllib3
from bs4 import BeautifulSoup
USER_AGENTS = [
    "Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; SV1; AcooBrowser; .NET CLR 1.1.4322; .NET CLR 2.0.50727)",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 6.0; Acoo Browser; SLCC1; .NET CLR 2.0.50727; Media Center PC 5.0; .NET CLR 3.0.04506)",
    "Mozilla/4.0 (compatible; MSIE 7.0; AOL 9.5; AOLBuild 4337.35; Windows NT 5.1; .NET CLR 1.1.4322; .NET CLR 2.0.50727)",
    "Mozilla/5.0 (Windows; U; MSIE 9.0; Windows NT 9.0; en-US)",
    "Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Win64; x64; Trident/5.0; .NET CLR 3.5.30729; .NET CLR 3.0.30729; .NET CLR 2.0.50727; Media Center PC 6.0)",
    "Mozilla/5.0 (compatible; MSIE 8.0; Windows NT 6.0; Trident/4.0; WOW64; Trident/4.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; .NET CLR 1.0.3705; .NET CLR 1.1.4322)",
    "Mozilla/4.0 (compatible; MSIE 7.0b; Windows NT 5.2; .NET CLR 1.1.4322; .NET CLR 2.0.50727; InfoPath.2; .NET CLR 3.0.04506.30)",
    "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN) AppleWebKit/523.15 (KHTML, like Gecko, Safari/419.3) Arora/0.3 (Change: 287 c9dfb30)",
    "Mozilla/5.0 (X11; U; Linux; en-US) AppleWebKit/527+ (KHTML, like Gecko, Safari/419.3) Arora/0.6",
    "Mozilla/5.0 (Windows; U; Windows NT 5.1; en-US; rv:1.8.1.2pre) Gecko/20070215 K-Ninja/2.1.1",
    "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN; rv:1.9) Gecko/20080705 Firefox/3.0 Kapiko/3.0",
    "Mozilla/5.0 (X11; Linux i686; U;) Gecko/20070322 Kazehakase/0.4.5",
    "Mozilla/5.0 (X11; U; Linux i686; en-US; rv:1.9.0.8) Gecko Fedora/1.9.0.8-1.fc10 Kazehakase/0.5.6",
    "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.56 Safari/535.11",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_7_3) AppleWebKit/535.20 (KHTML, like Gecko) Chrome/19.0.1036.7 Safari/535.20",
    "Opera/9.80 (Macintosh; Intel Mac OS X 10.6.8; U; fr) Presto/2.9.168 Version/11.52",
    "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.11 (KHTML, like Gecko) Chrome/20.0.1132.11 TaoBrowser/2.0 Safari/536.11",
    "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/21.0.1180.71 Safari/537.1 LBBROWSER",
    "Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; WOW64; Trident/5.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; .NET4.0C; .NET4.0E; LBBROWSER)",
    "Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; SV1; QQDownload 732; .NET4.0C; .NET4.0E; LBBROWSER)",
    "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.84 Safari/535.11 LBBROWSER",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 6.1; WOW64; Trident/5.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; .NET4.0C; .NET4.0E)",
    "Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; WOW64; Trident/5.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; .NET4.0C; .NET4.0E; QQBrowser/7.0.3698.400)",
    "Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; SV1; QQDownload 732; .NET4.0C; .NET4.0E)",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; Trident/4.0; SV1; QQDownload 732; .NET4.0C; .NET4.0E; 360SE)",
    "Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; SV1; QQDownload 732; .NET4.0C; .NET4.0E)",
    "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 6.1; WOW64; Trident/5.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; .NET4.0C; .NET4.0E)",
    "Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/21.0.1180.89 Safari/537.1",
    "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/21.0.1180.89 Safari/537.1",
    "Mozilla/5.0 (iPad; U; CPU OS 4_2_1 like Mac OS X; zh-cn) AppleWebKit/533.17.9 (KHTML, like Gecko) Version/5.0.2 Mobile/8C148 Safari/6533.18.5",
    "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:2.0b13pre) Gecko/20110307 Firefox/4.0b13pre",
    "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:16.0) Gecko/20100101 Firefox/16.0",
    "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.11 (KHTML, like Gecko) Chrome/23.0.1271.64 Safari/537.11",
    "Mozilla/5.0 (X11; U; Linux x86_64; zh-CN; rv:1.9.2.10) Gecko/20100922 Ubuntu/10.10 (maverick) Firefox/3.6.10"
]
headers = {
    'User-Agent': USER_AGENTS[random.randint(0,33)],
    'Content-Type': "application/x-www-form-urlencoded",
}

def corps():
    urls = 'https://www.butian.net/Reward/corps'
    payload = {
        's':'3',
        'p':'1',
        'sort':'1'
    }
    r = requests.post(urls, data=payload, headers=headers)
    r.encoding = 'utf-8'
    print(r.status_code)
    res = r.text
    # print(json.loads(res))
    # print(type(json.loads(res)))
    # print(json.loads(res).keys())
    data = json.loads(res)['data']
    print("| {0:{3}^20} | {1:^15} | {2:<10}".format('企业名称', '赏金', '企业SRC地址', chr(12288)))
    print("-----------------------------------------------------------")
    for i in data['list']:
        company_src = 'https://www.butian.net/Company/%s'%str(i['company_id'])
        print("| {0:{3}^20} | {1:^15} | {2:<10}".format(i['company_name'], i['reward'], company_src, chr(12288)))

def company():
    urls = 'https://www.butian.net/Company/'
    data = []
    for i in range(1,30):
        url = urls + str(i)
        try:
            res = requests.post(url, headers=headers, allow_redirects=False, verify=False)
        except requests.exceptions.ConnectionError:
            print("网络错误,连接中止.")
            exit()
        except requests.exceptions.HTTPError:
            continue
        if(res.status_code == 200):
            # print("{0}\t{1}".format(res.url, res.status_code))
            texts = BeautifulSoup(res.text, 'html.parser')
            name  = texts.find(class_='firmName1').get_text()   # 企业名称

            spp1, spp2 = texts.find_all(class_='spp')
            spp1, spp2 = spp1.text , spp2.text

            firmDec = texts.find(class_='firmDec')
            firmDec.dl.decompose()
            firmDec.dl.decompose()
            # ?-该处的解析未解决，无法按要求输出注释内容。
            firmDec = str(firmDec)

            list = {
                'id'  : i ,
                'name': name , # 企业名称
                'type': firmDec[-13:-11], # 公益/专属
                'spp1': spp1 , # 已发现漏洞数/最高赏金
                'spp2': spp2 , # 已解决漏洞数/累计赏金
                'urls': res.url , # src_url
            }
            # print(dict)
            data.append(list)
        else:
            continue
    # print(data)
    company = {
        'status': 'ok' ,
        'length': len(data) ,
        'data'  : data,
    }
    # print(company)
    # print(json.dumps(company))
    return json.dumps(company)

def companyOut(data):
    # print(data)
    Out = "| {0:^6} | {5:<40} | {2:{6}^4} | {3:^10} | {4:^10} | {1:<10}"
    wb = openpyxl.Workbook()
    ws = wb.active

    sheetTitleList = ['id','企业名称','类型','已发现/最高赏金','已解决/累计赏金','SRC地址']
    k = 1
    for titles in sheetTitleList:
        ws.cell(row=1, column=k, value=titles)
        k = k+1
    row = 2
    for i in data:
        ws.cell(row=row, column=1, value=i['id'])
        ws.cell(row=row, column=2, value=i['name'])
        ws.cell(row=row, column=3, value=i['type'])
        ws.cell(row=row, column=4, value=i['spp1'])
        ws.cell(row=row, column=5, value=i['spp2'])
        ws.cell(row=row, column=6, value=i['urls'])
        row = row+1

    wb.save('补天SRC名录表.xlsx')



def main():
    pass

if __name__ == '__main__':
    urllib3.disable_warnings()
    company_data = company()
    print("###### 数据爬取完成，共计%s条数据 ######"%(str(len(json.loads(company_data)['data']))))
    companyOut(json.loads(company_data)['data'])
    print("###### 数据保存完成 ######")
